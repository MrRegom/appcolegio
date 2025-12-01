"""
Service Layer para importacion de datos desde Excel.

Contiene la logica de negocio para importar mantenedores desde archivos Excel.
Sigue Clean Architecture: separacion de responsabilidades.
"""
from typing import List, Dict, Any, Tuple, Optional, Type
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from django.core.exceptions import ValidationError
from django.db import transaction, models
from io import BytesIO


class ImportacionExcelService:
    """
    Service para importacion de datos desde Excel.
    
    Proporciona metodos genericos para:
    - Generar plantillas Excel con datos reales
    - Validar archivos Excel
    - Importar datos desde Excel
    """
    
    @staticmethod
    def validar_archivo_excel(archivo) -> Tuple[bool, str]:
        """
        Valida que el archivo sea un Excel valido.
        
        Args:
            archivo: Archivo subido
            
        Returns:
            Tuple[bool, str]: (es_valido, mensaje_error)
        """
        if not archivo:
            return False, "No se proporciono archivo"
        
        if not archivo.name.endswith(('.xlsx', '.xls')):
            return False, "El archivo debe ser un Excel (.xlsx o .xls)"
        
        try:
            wb = load_workbook(archivo, read_only=True)
            wb.close()
            return True, ""
        except Exception as e:
            return False, f"Error al leer el archivo: {str(e)}"
    
    @staticmethod
    def leer_datos_desde_excel(archivo, columnas_esperadas: List[str], fila_inicio: int = 2) -> List[Dict[str, Any]]:
        """
        Lee datos desde un archivo Excel.
        
        Args:
            archivo: Archivo Excel subido
            columnas_esperadas: Lista de nombres de columnas esperadas
            fila_inicio: Fila donde comienzan los datos (default: 2, asumiendo fila 1 es encabezado)
            
        Returns:
            Lista de diccionarios con los datos leidos
        """
        wb = load_workbook(archivo, read_only=True)
        ws = wb.active
        
        datos = []
        
        # Leer encabezados de la primera fila
        encabezados = []
        for cell in ws[1]:
            encabezados.append(cell.value if cell.value else "")
        
        # Validar que las columnas esperadas esten presentes
        columnas_faltantes = [col for col in columnas_esperadas if col not in encabezados]
        if columnas_faltantes:
            wb.close()
            raise ValidationError(f"Columnas faltantes en el archivo: {', '.join(columnas_faltantes)}")
        
        # Leer datos desde la fila de inicio
        for row in ws.iter_rows(min_row=fila_inicio, values_only=False):
            # Saltar filas vacias
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            
            # Crear diccionario con los datos de la fila
            fila_data = {}
            for idx, header in enumerate(encabezados):
                if idx < len(row):
                    valor = row[idx].value
                    # Convertir None a string vacio
                    fila_data[header] = str(valor).strip() if valor is not None else ""
            
            datos.append(fila_data)
        
        wb.close()
        return datos
    
    @staticmethod
    def generar_plantilla_generica(
        modelo: Type[models.Model],
        nombre_hoja: str,
        campos: List[Dict[str, str]],
        limite_registros: int = 10
    ) -> bytes:
        """
        Genera una plantilla Excel con datos reales del modelo.
        
        Args:
            modelo: Modelo de Django
            nombre_hoja: Nombre de la hoja Excel
            campos: Lista de diccionarios con {'nombre_campo': 'Nombre Columna', 'tipo': 'text|bool|number'}
            limite_registros: Cantidad maxima de registros a incluir (default: 10)
            
        Returns:
            bytes: Contenido del archivo Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = nombre_hoja
        
        # Encabezados
        encabezados = [campo['columna'] for campo in campos]
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Obtener datos reales del modelo (maximo limite_registros)
        registros = modelo.objects.filter(eliminado=False).order_by('codigo')[:limite_registros]
        
        # Si hay datos, usarlos; si no, dejar vacio
        if registros.exists():
            for row_idx, registro in enumerate(registros, start=2):
                for col_idx, campo in enumerate(campos, start=1):
                    nombre_campo = campo['nombre_campo']
                    tipo = campo.get('tipo', 'text')
                    
                    # Obtener valor del campo
                    valor = getattr(registro, nombre_campo, '')
                    
                    # Convertir segun tipo
                    if tipo == 'bool':
                        valor = 'SI' if valor else 'NO'
                    elif valor is None:
                        valor = ''
                    
                    ws.cell(row=row_idx, column=col_idx, value=str(valor))
        
        # Ajustar ancho de columnas
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(encabezados) + 1):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = 20
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido
    
    @staticmethod
    def importar_generico(
        archivo,
        modelo: Type[models.Model],
        campos: List[Dict[str, str]],
        campo_clave: str = 'codigo',
        usuario=None
    ) -> Tuple[int, int, List[str]]:
        """
        Importa datos desde Excel de forma generica.
        
        Args:
            archivo: Archivo Excel
            modelo: Modelo de Django
            campos: Lista de diccionarios con {'nombre_campo': 'Nombre Columna', 'tipo': 'text|bool|number'}
            campo_clave: Campo unico para identificar registros (default: 'codigo')
            usuario: Usuario que realiza la importacion
            
        Returns:
            Tuple[int, int, List[str]]: (creadas, actualizadas, errores)
        """
        columnas_esperadas = [campo['columna'] for campo in campos]
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        # Crear mapeo de columnas a campos
        mapeo_columnas = {campo['columna']: campo for campo in campos}
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):
                try:
                    # Obtener valor del campo clave
                    campo_clave_col = next(
                        (c['columna'] for c in campos if c['nombre_campo'] == campo_clave),
                        None
                    )
                    
                    if not campo_clave_col:
                        errores.append(f"Fila {idx}: Campo clave '{campo_clave}' no encontrado")
                        continue
                    
                    valor_clave = fila.get(campo_clave_col, '').strip()
                    if not valor_clave:
                        errores.append(f"Fila {idx}: {campo_clave_col} es obligatorio")
                        continue
                    
                    # Construir defaults para update_or_create
                    defaults = {}
                    for campo in campos:
                        nombre_campo = campo['nombre_campo']
                        if nombre_campo == campo_clave:
                            continue
                        
                        columna = campo['columna']
                        tipo = campo.get('tipo', 'text')
                        valor = fila.get(columna, '').strip()
                        
                        if tipo == 'bool':
                            defaults[nombre_campo] = valor.upper() in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                        elif tipo == 'number':
                            try:
                                defaults[nombre_campo] = int(valor) if valor else 0
                            except ValueError:
                                defaults[nombre_campo] = 0
                        else:
                            defaults[nombre_campo] = valor if valor else ''
                    
                    # Siempre asegurar que eliminado sea False
                    defaults['eliminado'] = False
                    
                    # Buscar o crear
                    filtro = {campo_clave: valor_clave}
                    registro, created = modelo.objects.update_or_create(
                        **filtro,
                        defaults=defaults
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores
    
    # ==================== METODOS ESPECIFICOS POR MODELO ====================
    
    @staticmethod
    def generar_plantilla_marcas() -> bytes:
        """Genera plantilla de marcas con datos reales."""
        from apps.bodega.models import Marca
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.generar_plantilla_generica(Marca, 'Marcas', campos)
    
    @staticmethod
    def importar_marcas(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa marcas desde Excel."""
        from apps.bodega.models import Marca
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.importar_generico(archivo, Marca, campos, 'codigo', usuario)
    
    @staticmethod
    def generar_plantilla_operaciones() -> bytes:
        """Genera plantilla de operaciones con datos reales."""
        from apps.bodega.models import Operacion
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'tipo', 'columna': 'Tipo', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.generar_plantilla_generica(Operacion, 'Operaciones', campos)
    
    @staticmethod
    def importar_operaciones(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa operaciones desde Excel."""
        from apps.bodega.models import Operacion
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'tipo', 'columna': 'Tipo', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.importar_generico(archivo, Operacion, campos, 'codigo', usuario)
    
    @staticmethod
    def generar_plantilla_tipos_movimiento() -> bytes:
        """Genera plantilla de tipos de movimiento con datos reales."""
        from apps.bodega.models import TipoMovimiento
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.generar_plantilla_generica(TipoMovimiento, 'TiposMovimiento', campos)
    
    @staticmethod
    def importar_tipos_movimiento(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa tipos de movimiento desde Excel."""
        from apps.bodega.models import TipoMovimiento
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.importar_generico(archivo, TipoMovimiento, campos, 'codigo', usuario)
    
    @staticmethod
    def generar_plantilla_tipos_solicitud() -> bytes:
        """Genera plantilla de tipos de solicitud con datos reales."""
        from apps.solicitudes.models import TipoSolicitud
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'requiere_aprobacion', 'columna': 'RequiereAprobacion', 'tipo': 'bool'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.generar_plantilla_generica(TipoSolicitud, 'TiposSolicitud', campos)
    
    @staticmethod
    def importar_tipos_solicitud(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa tipos de solicitud desde Excel."""
        from apps.solicitudes.models import TipoSolicitud
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'requiere_aprobacion', 'columna': 'RequiereAprobacion', 'tipo': 'bool'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.importar_generico(archivo, TipoSolicitud, campos, 'codigo', usuario)
    
    @staticmethod
    def generar_plantilla_estados_solicitud() -> bytes:
        """Genera plantilla de estados de solicitud con datos reales."""
        from apps.solicitudes.models import EstadoSolicitud
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'color', 'columna': 'Color', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.generar_plantilla_generica(EstadoSolicitud, 'EstadosSolicitud', campos)
    
    @staticmethod
    def importar_estados_solicitud(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa estados de solicitud desde Excel."""
        from apps.solicitudes.models import EstadoSolicitud
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'color', 'columna': 'Color', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.importar_generico(archivo, EstadoSolicitud, campos, 'codigo', usuario)
    
    @staticmethod
    def generar_plantilla_estados_recepcion() -> bytes:
        """Genera plantilla de estados de recepcion con datos reales."""
        from apps.compras.models import EstadoRecepcion
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.generar_plantilla_generica(EstadoRecepcion, 'EstadosRecepcion', campos)
    
    @staticmethod
    def importar_estados_recepcion(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa estados de recepcion desde Excel."""
        from apps.compras.models import EstadoRecepcion
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.importar_generico(archivo, EstadoRecepcion, campos, 'codigo', usuario)
    
    @staticmethod
    def generar_plantilla_tipos_recepcion() -> bytes:
        """Genera plantilla de tipos de recepcion con datos reales."""
        from apps.compras.models import TipoRecepcion
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'requiere_orden', 'columna': 'RequiereOrden', 'tipo': 'bool'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.generar_plantilla_generica(TipoRecepcion, 'TiposRecepcion', campos)
    
    @staticmethod
    def importar_tipos_recepcion(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa tipos de recepcion desde Excel."""
        from apps.compras.models import TipoRecepcion
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'requiere_orden', 'columna': 'RequiereOrden', 'tipo': 'bool'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.importar_generico(archivo, TipoRecepcion, campos, 'codigo', usuario)
    
    @staticmethod
    def generar_plantilla_estados_orden_compra() -> bytes:
        """Genera plantilla de estados de orden de compra con datos reales."""
        from apps.compras.models import EstadoOrdenCompra
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'color', 'columna': 'Color', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.generar_plantilla_generica(EstadoOrdenCompra, 'EstadosOrdenCompra', campos)
    
    @staticmethod
    def importar_estados_orden_compra(archivo, usuario) -> Tuple[int, int, List[str]]:
        """Importa estados de orden de compra desde Excel."""
        from apps.compras.models import EstadoOrdenCompra
        campos = [
            {'nombre_campo': 'codigo', 'columna': 'Codigo', 'tipo': 'text'},
            {'nombre_campo': 'nombre', 'columna': 'Nombre', 'tipo': 'text'},
            {'nombre_campo': 'descripcion', 'columna': 'Descripcion', 'tipo': 'text'},
            {'nombre_campo': 'color', 'columna': 'Color', 'tipo': 'text'},
            {'nombre_campo': 'activo', 'columna': 'Activo', 'tipo': 'bool'},
        ]
        return ImportacionExcelService.importar_generico(archivo, EstadoOrdenCompra, campos, 'codigo', usuario)
