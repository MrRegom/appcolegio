"""
Service Layer para importacion de datos desde Excel.

Contiene la logica de negocio para importar mantenedores desde archivos Excel.
Sigue Clean Architecture: separacion de responsabilidades.
"""
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook
from django.core.exceptions import ValidationError
from django.db import transaction


class ImportacionExcelService:
    """
    Service para importacion de datos desde Excel.
    
    Proporciona metodos genericos para:
    - Generar plantillas Excel
    - Validar archivos Excel
    - Importar datos desde Excel
    """
    
    @staticmethod
    def validar_archivo_excel(archivo) -> Tuple[bool, str]:
        """
        Valida que el archivo sea un Excel valido.
        
        Args:
            archivo: Archivo subido
            
        Returns:
            Tuple[bool, str]: (es_valido, mensaje_error)
        """
        if not archivo:
            return False, "No se proporciono archivo"
        
        if not archivo.name.endswith(('.xlsx', '.xls')):
            return False, "El archivo debe ser un Excel (.xlsx o .xls)"
        
        try:
            wb = load_workbook(archivo, read_only=True)
            wb.close()
            return True, ""
        except Exception as e:
            return False, f"Error al leer el archivo: {str(e)}"
    
    @staticmethod
    def leer_datos_desde_excel(archivo, columnas_esperadas: List[str], fila_inicio: int = 2) -> List[Dict[str, Any]]:
        """
        Lee datos desde un archivo Excel.
        
        Args:
            archivo: Archivo Excel subido
            columnas_esperadas: Lista de nombres de columnas esperadas
            fila_inicio: Fila donde comienzan los datos (default: 2, asumiendo fila 1 es encabezado)
            
        Returns:
            Lista de diccionarios con los datos leidos
        """
        wb = load_workbook(archivo, read_only=True)
        ws = wb.active
        
        datos = []
        
        # Leer encabezados de la primera fila
        encabezados = []
        for cell in ws[1]:
            encabezados.append(cell.value if cell.value else "")
        
        # Validar que las columnas esperadas esten presentes
        columnas_faltantes = [col for col in columnas_esperadas if col not in encabezados]
        if columnas_faltantes:
            wb.close()
            raise ValidationError(f"Columnas faltantes en el archivo: {', '.join(columnas_faltantes)}")
        
        # Leer datos desde la fila de inicio
        for row in ws.iter_rows(min_row=fila_inicio, values_only=False):
            # Saltar filas vacias
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue
            
            # Crear diccionario con los datos de la fila
            fila_data = {}
            for idx, header in enumerate(encabezados):
                if idx < len(row):
                    valor = row[idx].value
                    # Convertir None a string vacio
                    fila_data[header] = str(valor).strip() if valor is not None else ""
            
            datos.append(fila_data)
        
        wb.close()
        return datos
    
    @staticmethod
    def importar_marcas(archivo, usuario) -> Tuple[int, int, List[str]]:
        """
        Importa marcas desde un archivo Excel.
        
        Args:
            archivo: Archivo Excel con las marcas
            usuario: Usuario que realiza la importacion
            
        Returns:
            Tuple[int, int, List[str]]: (creadas, actualizadas, errores)
        """
        from apps.bodega.models import Marca
        
        columnas_esperadas = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        datos = ImportacionExcelService.leer_datos_desde_excel(archivo, columnas_esperadas)
        
        creadas = 0
        actualizadas = 0
        errores = []
        
        with transaction.atomic():
            for idx, fila in enumerate(datos, start=2):  # Empezar desde fila 2 (despues del encabezado)
                try:
                    codigo = fila.get('Codigo', '').strip()
                    nombre = fila.get('Nombre', '').strip()
                    descripcion = fila.get('Descripcion', '').strip()
                    activo_str = fila.get('Activo', 'SI').strip().upper()
                    
                    if not codigo or not nombre:
                        errores.append(f"Fila {idx}: Codigo y Nombre son obligatorios")
                        continue
                    
                    activo = activo_str in ['SI', 'S', 'TRUE', '1', 'ACTIVO']
                    
                    marca, created = Marca.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'descripcion': descripcion,
                            'activo': activo,
                            'eliminado': False,
                        }
                    )
                    
                    if created:
                        creadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores.append(f"Fila {idx}: {str(e)}")
        
        return creadas, actualizadas, errores
    
    @staticmethod
    def generar_plantilla_marcas() -> bytes:
        """
        Genera una plantilla Excel para importar marcas.
        
        Returns:
            bytes: Contenido del archivo Excel
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Marcas"
        
        # Encabezados
        encabezados = ['Codigo', 'Nombre', 'Descripcion', 'Activo']
        for col_idx, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Ejemplo de datos
        ejemplos = [
            ['MAR-001', 'Marca Ejemplo 1', 'Descripcion de la marca ejemplo 1', 'SI'],
            ['MAR-002', 'Marca Ejemplo 2', 'Descripcion de la marca ejemplo 2', 'SI'],
        ]
        
        for row_idx, ejemplo in enumerate(ejemplos, start=2):
            for col_idx, valor in enumerate(ejemplo, start=1):
                ws.cell(row=row_idx, column=col_idx, value=valor)
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        contenido = output.read()
        output.close()
        
        return contenido

